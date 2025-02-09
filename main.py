from flask import Flask, render_template, request, jsonify
import re
import pandas as pd
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.mkdir(UPLOAD_FOLDER)


def converter_km_para_float(km):
    partes = km.split("+")
    if len(partes) == 2:
        return int(partes[0]) + int(partes[1]) / 1000
    return float(km)


def salvar_excel(dados_totais, aba, filename):
    arquivo_excel = filename

    if not os.path.exists(arquivo_excel):
        wb = Workbook()
        wb.save(arquivo_excel)

    wb = load_workbook(arquivo_excel)

    if aba not in wb.sheetnames:
        wb.create_sheet(title=aba)

    ws = wb[aba]

    if ws.max_row == 1 and not ws.cell(row=1, column=1).value:
        cabecalhos = ['Data', 'Atividade', 'Auditoria', 'Equipe', 'Regional',
                      'Rodovia', 'Km Inicial', 'Km Final', 'Extensão (Km)', 'Total de Elementos', 'Situação da Rodovia']
        ws.append(cabecalhos)

        for col in range(1, len(cabecalhos) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    for dado in dados_totais:
        km_inicio = converter_km_para_float(dado['Km Inicial'])
        km_final = converter_km_para_float(dado['Km Final'])
        extensao = round(km_final - km_inicio, 3)

        nova_linha = [
            dado['Data'], dado['Atividade'], dado['Auditoria'], dado['Equipe'],
            dado['Regional'], dado['Rodovia'], dado['Km Inicial'], dado['Km Final'],
            extensao,
            dado['Total de Elementos'], dado['Situação da Rodovia']
        ]
        ws.append(nova_linha)

        for col in range(1, len(nova_linha) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(arquivo_excel)
    wb.close()

    print("\nInformações salvas com sucesso no arquivo 'diario_de_campo_atualizado.xlsx'!")


import re

def processar_mensagem(mensagem, atividade):
    data = re.search(r"Data:\s*(\d{2}/\d{2}/\d{4})", mensagem)
    auditoria = re.search(r"Auditoria:\s*(\d+)", mensagem)
    equipe = re.search(r"Equipe:\s*(.+)", mensagem)
    regional = re.search(r"Regional:\s*(.+)", mensagem)

    dados_totais = []

    print("Mensagem recebida:\n", mensagem)

    rodovia_pattern = re.findall(
        r"(SP[A]?\d{3}(?:/\d{3})?)\s*do\s*Km\s*(\d{3}\+\d{3})\s*ao\s*Km\s*(\d{3}\+\d{3})\s*"
        r"Total de elementos:\s*(\d+)\s*"
        r"Situação:\s*([^\n\r]+)",
        mensagem,
        re.MULTILINE
    )

    print("Rodovias encontradas:", rodovia_pattern)

    if not rodovia_pattern:
        print("⚠️ Nenhuma rodovia encontrada! Verifique a formatação da mensagem.")
        return []

    for rodovia in rodovia_pattern:
        dados_totais.append({
            'Data': data.group(1) if data else '',
            'Atividade': atividade,
            'Auditoria': int(auditoria.group(1)) if auditoria else 0,
            'Equipe': equipe.group(1).strip() if equipe else '',
            'Regional': regional.group(1).strip() if regional else '',
            'Rodovia': rodovia[0],
            'Km Inicial': rodovia[1],
            'Km Final': rodovia[2],
            'Total de Elementos': int(rodovia[3]),
            'Situação da Rodovia': rodovia[4].strip()
        })

    print("Dados processados:\n", dados_totais)
    return dados_totais



app = Flask(__name__)


@app.route('/')
def home():
    return render_template('index.html')


@app.route('/api/verificar', methods=['POST'])
def api_view():
    file = request.files.get('file')

    if not file:
        return jsonify({"bool": False, "error": "Nenhum arquivo enviado", 'id_input': 1}), 400

    message = request.form.get('message')

    if not message:
        return jsonify({"bool": False, "error": "Mensagem não enviada", 'id_input': 2}), 400

    aba = request.form.get('type')

    if not aba:
        return jsonify({"bool": False, "error": "Nenhuma aba enviada", 'id_input': 3}), 400

    dados = processar_mensagem(message, aba)

    return jsonify({
        'bool': True,
        'sucesso': 'Arquivo e mensagens processados com sucesso!',
        'dados': dados
    })


@app.route('/api/adicionar', methods=['POST'])
def api_add():
    file = request.files.get('file')

    if not file:
        return jsonify({"bool": False, "error": "Nenhum arquivo enviado", 'id_input': 1}), 400

    message = request.form.get('message')

    if not message:
        return jsonify({"bool": False, "error": "Mensagem não enviada", 'id_input': 2}), 400

    aba = request.form.get('type')

    if not aba:
        return jsonify({"bool": False, "error": "Nenhuma aba enviada", 'id_input': 3}), 400

    folder_path = './uploads'
    os.makedirs(folder_path, exist_ok=True)

    file_path = os.path.join(folder_path, file.filename)
    file.save(file_path)

    dados = processar_mensagem(message, aba)

    salvar_excel(dados, aba, file.filename)

    return jsonify({"bool": True, "sucesso": "Arquivo e mensagem processados com sucesso!"})


@app.errorhandler(404)
def not_found(error):
    return "Página não encontrada!", 404


if __name__ == '__main__':
    app.run(debug=True)
